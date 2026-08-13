from openpyxl import load_workbook
from openpyxl.cell import MergedCell
import io
from .config import get_settings

settings = get_settings()

CELL_MAP = {
    "proponente": (4, 2),
    "razao_social": (5, 2),
    "cnpj": (6, 2),
    "rua": (7, 2),
    "numero": (8, 2),
    "complemento": (9, 2),
    "bairro": (10, 2),
    "cidade": (11, 2),
    "estado": (12, 2),
    "email": (13, 2),
    "redes_sociais": (14, 2),
    "rep_nome": (16, 2),
    "rep_cargo": (17, 2),
    "rep_documento": (18, 2),
    "rep_telefone": (19, 2),
    "rep_celular": (20, 2),
    "rep_email": (21, 2),
    "lei_incentivo": (23, 2),
    "nome_projeto": (24, 2),
    "codigo_aprovacao": (25, 2),
    "artigo_aprovacao": (26, 2),
    "data_publicacao": (27, 2),
    "data_prorrogacao": (28, 2),
    "data_final_captacao": (29, 2),
    "valor_total_aprovacao": (30, 2),
    "valor_solicitado_aporte": (31, 2),
    "local_realizacao": (32, 2),
    "periodo_realizacao": (33, 2),
    "objetivo": (34, 2),
    "publico_alvo": (35, 2),
    "detalhamento_atividades": (36, 2),
    "cronograma": (37, 2),
    "banco": (40, 2),
    "agencia": (41, 2),
    "conta_corrente": (42, 2),
    "operacao": (43, 2),
    "valores_organizacionais_nota": (47, 3),
    "valores_organizacionais_obs": (47, 4),
    "diversidade_inclusao_nota": (48, 3),
    "diversidade_inclusao_obs": (48, 4),
    "sustentabilidade_nota": (49, 3),
    "sustentabilidade_obs": (49, 4),
    "portfolio_nota": (55, 3),
    "portfolio_obs": (55, 4),
    "experiencia_incentivos_nota": (56, 3),
    "experiencia_incentivos_obs": (56, 4),
    "capacidade_tecnica_nota": (57, 3),
    "capacidade_tecnica_obs": (57, 4),
    "governanca_nota": (58, 3),
    "governanca_obs": (58, 4),
    "recursos_humanos_nota": (59, 3),
    "recursos_humanos_obs": (59, 4),
    "recursos_financeiros_nota": (60, 3),
    "recursos_financeiros_obs": (60, 4),
    "experiencia_resultados_nota": (61, 3),
    "experiencia_resultados_obs": (61, 4),
    "parcerias_nota": (62, 3),
    "parcerias_obs": (62, 4),
    "beneficiarios_diretos_nota": (67, 3),
    "beneficiarios_diretos_obs": (67, 4),
    "beneficiarios_indiretos_nota": (68, 3),
    "beneficiarios_indiretos_obs": (68, 4),
    "educacao_nota": (69, 3),
    "educacao_obs": (69, 4),
    "saude_nota": (70, 3),
    "saude_obs": (70, 4),
    "inclusao_nota": (71, 3),
    "inclusao_obs": (71, 4),
    "esg_nota": (72, 3),
    "esg_obs": (72, 4),
    "diferencial_artistico_nota": (73, 3),
    "diferencial_artistico_obs": (73, 4),
    "diferencial_social_nota": (74, 3),
    "diferencial_social_obs": (74, 4),
    "diferencial_originalidade_nota": (75, 3),
    "diferencial_originalidade_obs": (75, 4),
    "diferencial_tecnico_nota": (76, 3),
    "diferencial_tecnico_obs": (76, 4),
    "diferencial_relacionamento_nota": (77, 3),
    "diferencial_relacionamento_obs": (77, 4),
    "interesse_coletivo_nota": (78, 3),
    "interesse_coletivo_obs": (78, 4),
    "plano_comunicacao_nota": (83, 3),
    "plano_comunicacao_obs": (83, 4),
    "redes_sociais_nota": (84, 3),
    "redes_sociais_obs": (84, 4),
    "monitoramento_nota": (85, 3),
    "monitoramento_obs": (85, 4),
    "conteudo_institucional_nota": (86, 3),
    "conteudo_institucional_obs": (86, 4),
    "ativacoes_marca_nota": (87, 3),
    "ativacoes_marca_obs": (87, 4),
    "direitos_imagem_nota": (88, 3),
    "direitos_imagem_obs": (88, 4),
    "contrapartida_imagem_nota": (89, 3),
    "contrapartida_imagem_obs": (89, 4),
    "site_oficial_nota": (90, 3),
    "site_oficial_obs": (90, 4),
    "exibicao_video_nota": (91, 3),
    "exibicao_video_obs": (91, 4),
    "citacao_releases_nota": (92, 3),
    "citacao_releases_obs": (92, 4),
    "voluntariado_corporativo_nota": (97, 3),
    "voluntariado_corporativo_obs": (97, 4),
    "datas_comemorativas_nota": (98, 3),
    "datas_comemorativas_obs": (98, 4),
    "engajamento_comunitario_nota": (99, 3),
    "engajamento_comunitario_obs": (99, 4),
    "inclusao_material": (104, 4),
    "cessao_convites": (105, 4),
    "estande": (106, 4),
    "palestrantes": (107, 4),
    "cortesias_escolas": (108, 4),
    "citacao_evento": (109, 4),
    "captacao_nota": (114, 3),
    "captacao_obs": (114, 4),
    "execucao_garantida_nota": (115, 3),
    "execucao_garantida_obs": (115, 4),
    "cotas_nota": (116, 3),
    "cotas_obs": (116, 4),
}

def parse_val(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str) and v.strip() != "":
        try:
            return float(v.replace(",", ".").strip())
        except ValueError:
            return 0.0
    return 0.0

def _to_num_or_none(valor):
    """Tenta converter valor para número. Retorna None se for texto puro (descrição)."""
    if valor is None:
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    if isinstance(valor, str):
        v = valor.replace(",", ".").strip()
        try:
            return float(v)
        except ValueError:
            return None  # É um texto descritivo — vai para a coluna Obs
    return None

def gerar_planilha(dados: dict) -> bytes:
    wb = load_workbook(settings.MODELO_PLANILHA)
    ws = wb.active

    # Separar campos _nota que contêm texto (devem ir para Obs) dos que são numéricos
    for campo, valor in dados.items():
        if campo not in CELL_MAP or valor is None:
            continue

        linha, coluna = CELL_MAP[campo]

        if campo.endswith("_nota"):
            num = _to_num_or_none(valor)
            if num is not None:
                # Valor numérico → coluna Nota (col 3)
                cell_nota = ws.cell(row=linha, column=coluna)
                if not isinstance(cell_nota, MergedCell):
                    ws.cell(row=linha, column=coluna, value=num)
            else:
                # Texto descritivo → coluna Obs (col 4), não substituir Nota com texto
                cell_obs = ws.cell(row=linha, column=coluna + 1)
                if not isinstance(cell_obs, MergedCell):
                    ws.cell(row=linha, column=coluna + 1, value=str(valor))
        elif campo.endswith("_obs"):
            # Observação explícita → sempre vai para coluna 4
            cell = ws.cell(row=linha, column=coluna)
            if not isinstance(cell, MergedCell):
                ws.cell(row=linha, column=coluna, value=valor)
        else:
            # Campo de dados normais
            cell = ws.cell(row=linha, column=coluna)
            if not isinstance(cell, MergedCell):
                ws.cell(row=linha, column=coluna, value=valor)

    # Calcular subtotais por seção
    sec1_keys = ["valores_organizacionais_nota", "diversidade_inclusao_nota", "sustentabilidade_nota"]
    sec2_keys = ["portfolio_nota", "experiencia_incentivos_nota", "capacidade_tecnica_nota", "governanca_nota", "recursos_humanos_nota", "recursos_financeiros_nota", "experiencia_resultados_nota", "parcerias_nota"]
    sec3_keys = ["beneficiarios_diretos_nota", "beneficiarios_indiretos_nota", "educacao_nota", "saude_nota", "inclusao_nota", "esg_nota", "diferencial_artistico_nota", "diferencial_social_nota", "diferencial_originalidade_nota", "diferencial_tecnico_nota", "diferencial_relacionamento_nota", "interesse_coletivo_nota"]
    sec4_keys = ["plano_comunicacao_nota", "redes_sociais_nota", "monitoramento_nota", "conteudo_institucional_nota", "ativacoes_marca_nota", "direitos_imagem_nota", "contrapartida_imagem_nota", "site_oficial_nota", "exibicao_video_nota", "citacao_releases_nota"]
    sec5_keys = ["voluntariado_corporativo_nota", "datas_comemorativas_nota", "engajamento_comunitario_nota"]
    sec6_keys = ["captacao_nota", "execucao_garantida_nota", "cotas_nota"]

    def get_nota_num(k):
        v = dados.get(k)
        n = _to_num_or_none(v)
        return n if n is not None else 0.0

    s1 = sum(get_nota_num(k) for k in sec1_keys)
    s2 = sum(get_nota_num(k) for k in sec2_keys)
    s3 = sum(get_nota_num(k) for k in sec3_keys)
    s4 = sum(get_nota_num(k) for k in sec4_keys)
    s5 = sum(get_nota_num(k) for k in sec5_keys)
    s6 = sum(get_nota_num(k) for k in sec6_keys)

    ws.cell(row=124, column=2, value=s1)
    ws.cell(row=125, column=2, value=s2)
    ws.cell(row=126, column=2, value=s3)
    ws.cell(row=127, column=2, value=s4)
    ws.cell(row=128, column=2, value=s5)
    ws.cell(row=129, column=2, value=s6)

    ws.cell(row=79, column=3, value='=SUM(C67:C78)')
    ws.cell(row=93, column=3, value='=SUM(C83:C92)')

    # Escrever resumo do avaliador (IA) no painel executivo da planilha
    resumo = dados.get("resumo_avaliador") or dados.get("resumo_executivo")
    if resumo:
        ws.cell(row=133, column=1, value="Parecer da IA")
        cell_resumo = ws.cell(row=133, column=2)
        if not isinstance(cell_resumo, MergedCell):
            ws.cell(row=133, column=2, value=str(resumo))

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()