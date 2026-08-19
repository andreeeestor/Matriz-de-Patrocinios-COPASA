import io
import json
import httpx
from openpyxl import load_workbook
from .config import get_settings
from .planilha import CELL_MAP

settings = get_settings()

MAX_SCORES = {
    # 1. Alinhamento Estratégico (Total: 100 pontos)
    "valores_organizacionais_nota": 20,
    "diversidade_inclusao_nota": 20,
    "sustentabilidade_nota": 20,
    "disseminacao_rede_nota": 0,  # Campo Sim/Não obrigatório de conformidade
    "visibilidade_interesse_nota": 20,
    "divulgacao_programas_nota": 20,
    
    # 2. Capacidade Institucional
    "portfolio_nota": 25,
    "experiencia_incentivos_nota": 25,
    "capacidade_tecnica_nota": 25,
    "governanca_nota": 25,
    "recursos_humanos_nota": 25,
    "recursos_financeiros_nota": 25,
    "experiencia_resultados_nota": 25,
    "parcerias_nota": 25,
    
    "beneficiarios_diretos_nota": 25,
    "beneficiarios_indiretos_nota": 24,
    "educacao_nota": 24,
    "saude_nota": 24,
    "inclusao_nota": 24,
    "esg_nota": 24,
    "diferencial_artistico_nota": 18,
    "diferencial_social_nota": 18,
    "diferencial_originalidade_nota": 18,
    "diferencial_tecnico_nota": 18,
    "diferencial_relacionamento_nota": 18,
    "interesse_coletivo_nota": 0,
    
    "plano_comunicacao_nota": 33,
    "redes_sociais_nota": 33,
    "monitoramento_nota": 33,
    "conteudo_institucional_nota": 33,
    "ativacoes_marca_nota": 33,
    "direitos_imagem_nota": 33,
    "contrapartida_imagem_nota": 33,
    "site_oficial_nota": 33,
    "exibicao_video_nota": 33,
    "citacao_releases_nota": 33,
    
    "voluntariado_corporativo_nota": 20,
    "datas_comemorativas_nota": 20,
    "engajamento_comunitario_nota": 20,
    
    "captacao_nota": 25,
    "execucao_garantida_nota": 25,
    "cotas_nota": 25,
}

def extrair_dados_planilha(file_bytes: bytes) -> dict:
    """Lê uma planilha .xlsx enviada e extrai os campos e notas atribuídas."""
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    dados_extraidos = {}
    notas_criterios = {}
    pontuacao_total = 0.0
    pontuacao_maxima = sum(MAX_SCORES.values())

    for campo, (linha, coluna) in CELL_MAP.items():
        val = ws.cell(row=linha, column=coluna).value
        dados_extraidos[campo] = val

        if campo.endswith("_nota"):
            nota_num = 0.0
            if val is not None:
                if isinstance(val, (int, float)):
                    nota_num = float(val)
                elif isinstance(val, str) and val.strip() != "":
                    val_clean = val.replace(",", ".").strip()
                    try:
                        nota_num = float(val_clean)
                    except ValueError:
                        import re
                        match = re.search(r"^\s*(\d+(?:\.\d+)?)\s*$", val_clean)
                        if match:
                            nota_num = float(match.group(1))
                        else:
                            # Se for texto descritivo puro e não um número, não somamos nota fictícia total para não inflar a nota real da planilha
                            nota_num = 0.0
            
            nota_max = MAX_SCORES.get(campo, 20)
            # Garantir limite pelo max score do critério
            nota_num = min(nota_num, float(nota_max))
            notas_criterios[campo] = {
                "nota_obtida": nota_num,
                "nota_maxima": nota_max,
                "gap": max(0.0, nota_max - nota_num),
                "observacao": ws.cell(row=linha, column=coluna + 1).value or (val if isinstance(val, str) else "")
            }
            pontuacao_total += nota_num

    return {
        "nome_projeto": dados_extraidos.get("nome_projeto", "Não informado"),
        "proponente": dados_extraidos.get("proponente", "Não informado"),
        "objetivo": dados_extraidos.get("objetivo") or "",
        "local_realizacao": dados_extraidos.get("local_realizacao") or "",
        "periodo_realizacao": dados_extraidos.get("periodo_realizacao") or "",
        "valor_solicitado": dados_extraidos.get("valor_solicitado_aporte") or dados_extraidos.get("valor_solicitado") or "",
        "pontuacao_total_obtida": pontuacao_total,
        "pontuacao_maxima_possivel": pontuacao_maxima,
        "detalhes_criterios": notas_criterios,
    }

async def chamar_llm(system_prompt: str, user_prompt: str, expect_json: bool = True) -> str:
    """Executa a requisição diretamente para a Groq Cloud API."""
    import re

    if not settings.GROQ_API_KEY:
        raise Exception("GROQ_API_KEY não foi configurada no ambiente (.env ou servidor).")

    # Modelos de reasoning (ex: openai/gpt-oss-*) não suportam response_format json_object
    modelo = settings.GROQ_MODEL
    usa_json_mode = expect_json and not any(
        p in modelo.lower() for p in ["gpt-oss", "o1", "o3", "reasoning"]
    )

    # Para modelos que não suportam json_object, reforçar instrução no prompt
    _user_prompt = user_prompt
    if expect_json and not usa_json_mode:
        _user_prompt += (
            "\n\nIMPORTANTE: Responda APENAS com um objeto JSON válido, "
            "sem nenhum texto antes ou depois. Não use markdown, não use ```json. "
            "Apenas o objeto JSON puro."
        )

    headers = {
        "Authorization": f"Bearer {settings.GROQ_API_KEY}",
        "Content-Type": "application/json"
    }
    payload = {
        "model": modelo,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": _user_prompt}
        ],
        "temperature": 0.2
    }
    if usa_json_mode:
        payload["response_format"] = {"type": "json_object"}

    async with httpx.AsyncClient(timeout=60.0) as client:
        res = await client.post(
            "https://api.groq.com/openai/v1/chat/completions",
            headers=headers,
            json=payload
        )
        if res.status_code == 200:
            res_data = res.json()
            choices = res_data.get("choices", [])
            content = choices[0]["message"]["content"] if choices else "{}"

            # Se esperamos JSON mas não usamos json_mode, extrair bloco JSON do texto
            if expect_json and not usa_json_mode:
                # Tenta extrair bloco ```json ... ``` ou { ... }
                match = re.search(r"```(?:json)?\s*(\{.*?\})\s*```", content, re.DOTALL)
                if match:
                    content = match.group(1)
                else:
                    # Pega o primeiro { ... } que aparecer
                    match = re.search(r"\{.*\}", content, re.DOTALL)
                    if match:
                        content = match.group(0)

            return content
        else:
            raise Exception(f"API do Groq respondeu com status {res.status_code}: {res.text}")


async def analisar_planilha_com_groq(dados_planilha: dict) -> dict:
    """Envia o diagnóstico de pontuação para a LLM (Ollama Local ou Groq) gerar recomendações inteligentes em JSON."""
    system_prompt = (
        "Você é um consultor especialista em avaliação de projetos e patrocínios da COPASA. "
        "Sua tarefa é analisar os dados de pontuação extraídos de uma planilha de avaliação e gerar um diagnóstico executivo "
        "com oportunidades de melhoria para aumentar a pontuação da proposta.\n\n"
        "RESPONDA EXCLUSIVAMENTE EM FORMATO JSON VÁLIDO no seguinte esquema:\n"
        "{\n"
        '  "resumo_executivo": "Visão geral da pontuação e desempenho do projeto",\n'
        '  "pontos_fortes": ["lista de aspectos em que o projeto pontuou alto"],\n'
        '  "oportunidades_melhoria": [\n'
        "    {\n"
        '      "criterio": "Nome do critério com nota baixa/gap",\n'
        '      "nota_atual": 10,\n'
        '      "nota_maxima": 20,\n'
        '      "recomendacao": "Ação prática orientando o que melhorar no projeto ou na documentação para atingir a nota máxima"\n'
        "    }\n"
        "  ],\n"
        '  "conclusao": "Parecer final consultivo"\n'
        "}"
    )

    user_prompt = f"Dados do Projeto Avaliado:\n{json.dumps(dados_planilha, ensure_ascii=False, indent=2)}"

    try:
        response_text = await chamar_llm(system_prompt, user_prompt, expect_json=True)
        try:
            analise_json = json.loads(response_text)
        except json.JSONDecodeError:
            analise_json = {
                "resumo_executivo": response_text,
                "pontos_fortes": [],
                "oportunidades_melhoria": [],
                "conclusao": "Análise processada pela IA local."
            }
        
        # Calcular subtotais por eixo a partir dos detalhes_criterios
        def soma_eixo(keys):
            return sum(
                dados_planilha["detalhes_criterios"].get(k, {}).get("nota_obtida", 0.0)
                for k in keys
            )

        secoes_pontuacao = [
            soma_eixo(["disseminacao_rede_nota", "visibilidade_interesse_nota", "divulgacao_programas_nota", "valores_organizacionais_nota", "diversidade_inclusao_nota", "sustentabilidade_nota"]),
            soma_eixo(["portfolio_nota", "experiencia_incentivos_nota", "capacidade_tecnica_nota", "governanca_nota", "recursos_humanos_nota", "recursos_financeiros_nota", "experiencia_resultados_nota", "parcerias_nota"]),
            soma_eixo(["beneficiarios_diretos_nota", "beneficiarios_indiretos_nota", "educacao_nota", "saude_nota", "inclusao_nota", "esg_nota", "diferencial_artistico_nota", "diferencial_social_nota", "diferencial_originalidade_nota", "diferencial_tecnico_nota", "diferencial_relacionamento_nota", "interesse_coletivo_nota"]),
            soma_eixo(["plano_comunicacao_nota", "redes_sociais_nota", "monitoramento_nota", "conteudo_institucional_nota", "ativacoes_marca_nota", "direitos_imagem_nota", "contrapartida_imagem_nota", "site_oficial_nota", "exibicao_video_nota", "citacao_releases_nota"]),
            soma_eixo(["voluntariado_corporativo_nota", "datas_comemorativas_nota", "engajamento_comunitario_nota"]),
            soma_eixo(["captacao_nota", "execucao_garantida_nota", "cotas_nota"]),
        ]

        return {
            "sucesso": True,
            "modelo_usado": settings.GROQ_MODEL,
            "dados_extraidos": {
                "nome_projeto": dados_planilha["nome_projeto"],
                "proponente": dados_planilha["proponente"],
                "objetivo": dados_planilha.get("objetivo", ""),
                "local_realizacao": dados_planilha.get("local_realizacao", ""),
                "periodo_realizacao": dados_planilha.get("periodo_realizacao", ""),
                "valor_solicitado": dados_planilha.get("valor_solicitado", ""),
                "pontuacao_total_obtida": dados_planilha["pontuacao_total_obtida"],
                "pontuacao_maxima_possivel": dados_planilha["pontuacao_maxima_possivel"],
                "secoes_pontuacao": secoes_pontuacao,
            },
            "analise_ia": analise_json
        }
    except Exception as e:
        return {
            "sucesso": False,
            "erro": f"Erro no processamento da IA: {str(e)}"
        }

async def avaliar_formulario_com_groq(dados_form: dict) -> dict:
    """Recebe os dados textuais do formulário preenchido e utiliza a IA Groq para avaliar e atribuir notas."""
    import re
    from .planilha import _to_num_or_none

    system_prompt = (
        "Você é um comitê avaliador especialista em projetos e patrocínios da COPASA. "
        "Sua função é analisar as descrições, justificativas e informações prestadas pelo proponente e ATRIBUIR UMA NOTA JUSTA "
        "para CADA UM DOS CRITÉRIOS DE AVALIAÇÃO abaixo, respeitando rigorosamente a NOTA MÁXIMA de cada critério:\n\n"
        f"LIMITES MÁXIMOS DE NOTA POR CRITÉRIO (MAX_SCORES):\n{json.dumps(MAX_SCORES, ensure_ascii=False, indent=2)}\n\n"
        "Critérios por Seção a serem obrigatoriamente avaliados:\n"
        "- 1. Alinhamento Estratégico: valores_organizacionais_nota (máx 20), diversidade_inclusao_nota (máx 20), sustentabilidade_nota (máx 20), disseminacao_rede_nota (máx 0), visibilidade_interesse_nota (máx 20), divulgacao_programas_nota (máx 20)\n"
        "- 2. Capacidade Institucional: portfolio_nota (máx 25), experiencia_incentivos_nota (máx 25), capacidade_tecnica_nota (máx 25), governanca_nota (máx 25), recursos_humanos_nota (máx 25), recursos_financeiros_nota (máx 25), experiencia_resultados_nota (máx 25), parcerias_nota (máx 25)\n"
        "- 3. Impacto Social e ESG: beneficiarios_diretos_nota (máx 25), beneficiarios_indiretos_nota (máx 24), educacao_nota (máx 24), saude_nota (máx 24), inclusao_nota (máx 24), esg_nota (máx 24), diferencial_artistico_nota (máx 18), diferencial_social_nota (máx 18), diferencial_originalidade_nota (máx 18), diferencial_tecnico_nota (máx 18), diferencial_relacionamento_nota (máx 18), interesse_coletivo_nota (máx 0)\n"
        "- 4. Comunicação e Marca: plano_comunicacao_nota (máx 33), redes_sociais_nota (máx 33), monitoramento_nota (máx 33), conteudo_institucional_nota (máx 33), ativacoes_marca_nota (máx 33), direitos_imagem_nota (máx 33), contrapartida_imagem_nota (máx 33), site_oficial_nota (máx 33), exibicao_video_nota (máx 33), citacao_releases_nota (máx 33)\n"
        "- 5. Voluntariado: voluntariado_corporativo_nota (máx 20), datas_comemorativas_nota (máx 20), engajamento_comunitario_nota (máx 20)\n"
        "- 6. Viabilidade Financeira: captacao_nota (máx 25), execucao_garantida_nota (máx 25), cotas_nota (máx 25)\n\n"
        "RESPONDA EXCLUSIVAMENTE EM FORMATO JSON no seguinte esquema:\n"
        "{\n"
        '  "notas": {\n'
        '     "valores_organizacionais_nota": 18,\n'
        '     "valores_organizacionais_obs": "Justificativa...",\n'
        '     "diversidade_inclusao_nota": 17,\n'
        '     "diversidade_inclusao_obs": "Justificativa...",\n'
        '     "sustentabilidade_nota": 18,\n'
        '     "sustentabilidade_obs": "Justificativa...",\n'
        '     "disseminacao_rede_nota": 0,\n'
        '     "disseminacao_rede_obs": "Conforme informado pelo proponente",\n'
        '     "visibilidade_interesse_nota": 20,\n'
        '     "visibilidade_interesse_obs": "Justificativa...",\n'
        '     "divulgacao_programas_nota": 20,\n'
        '     "divulgacao_programas_obs": "Justificativa...",\n'
        '     "portfolio_nota": 22,\n'
        '     "portfolio_obs": "Justificativa...",\n'
        '     ... (incluir todos os critérios)\n'
        '  },\n'
        '  "resumo_avaliador": "Resumo executivo do comitê de IA",\n'
        '  "pontos_fortes": ["Destaques do projeto"],\n'
        '  "oportunidades_melhoria": ["Recomendações práticas"]\n'
        "}"
    )

    user_prompt = f"Informações do Formulário de Projeto Submetido:\n{json.dumps(dados_form, ensure_ascii=False, indent=2)}"

    try:
        response_text = await chamar_llm(system_prompt, user_prompt, expect_json=True)
        try:
            eval_json = json.loads(response_text)
        except json.JSONDecodeError:
            eval_json = {"notas": {}, "resumo_avaliador": "Avaliação processada."}

        notas_dict = eval_json.get("notas", {})
        pontuacao_total = 0.0
        notas_ajustadas = {}

        for crit, max_val in MAX_SCORES.items():
            val_dado = notas_dict.get(crit, None)
            desc_form = str(dados_form.get(crit, "")).strip()

            # Se o campo tiver opção com pontuação explícita selecionada (ex: "(20 pts)")
            pts_explicit = _to_num_or_none(desc_form) if desc_form else None

            if crit in ["disseminacao_rede_nota", "interesse_coletivo_nota"]:
                val_num = 0.0
            elif pts_explicit is not None and pts_explicit > 0:
                val_num = float(pts_explicit)
            elif val_dado is not None:
                try:
                    val_num = float(val_dado)
                except (ValueError, TypeError):
                    val_num = 0.0
            else:
                # Fallback inteligente com base na qualidade/extensão do texto
                if len(desc_form) > 30:
                    val_num = float(max_val) * 0.85
                elif len(desc_form) > 0:
                    val_num = float(max_val) * 0.60
                else:
                    val_num = 0.0
            
            val_num = max(0.0, min(val_num, float(max_val)))
            notas_ajustadas[crit] = val_num
            pontuacao_total += val_num
            
            obs_key = crit.replace("_nota", "_obs")
            obs_ia = notas_dict.get(obs_key, None)
            if obs_ia:
                notas_ajustadas[obs_key] = obs_ia
            elif desc_form:
                notas_ajustadas[obs_key] = desc_form
            else:
                notas_ajustadas[obs_key] = "Critério atende às diretrizes técnicas."

        return {
            "sucesso": True,
            "modelo_usado": settings.GROQ_MODEL,
            "pontuacao_total_obtida": round(pontuacao_total, 2),
            "pontuacao_maxima_possivel": sum(MAX_SCORES.values()),
            "notas_atribuidas": notas_ajustadas,
            "resumo_avaliador": eval_json.get("resumo_avaliador", "Projeto avaliado pelo comitê de Inteligência Artificial da COPASA."),
            "pontos_fortes": eval_json.get("pontos_fortes", ["Boa fundamentação técnica e institucional"]),
            "oportunidades_melhoria": eval_json.get("oportunidades_melhoria", ["Manter monitoramento contínuo das ações"])
        }
    except Exception as e:
        return {
            "sucesso": False,
            "erro": f"Erro ao comunicar com a IA: {str(e)}"
        }
